from __future__ import print_function
# Importing Necessary Modules
from googleapiclient.discovery import build
from httplib2 import Http
from oauth2client import file, client, tools
from datetime import date
import requests
from datetime import datetime
from googleapiclient.errors import HttpError
import os.path
from openpyxl import Workbook
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow


start_time = datetime.utcnow()
URL = 'https://sheetdb.io/api/v1/boquxjkqizfuy'
# Scopes of the API as defined by google
SCOPES = 'https://www.googleapis.com/auth/gmail.readonly'

def main():
    
    creds = None

    # Initializing Excel Sheet
    row_no = 2
    book = Workbook()
    sheet = book.active
    try:
        sheet.cell(row=1, column=1).value = "DATE"
        sheet.cell(row=1, column=2).value = "MORNING"
        sheet.cell(row=1, column=3).value = "AFTERNOON"
        sheet.cell(row=1, column=4).value = "NIGHT"
        book.save("funds2.xlsx")

    
    except :
        pass

    # Storing the tokens and credentials.
    # token.json contains authorised token of user
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
   
   
    service = build('gmail', 'v1', credentials=creds)
    

    breakfast = []
    lunch = []
    dinner = []
    lunch_expense = 0
    dinner_expense = 0 
    breakfast_expense = 0
    dates=[]
    # sender is the user id of the service providing transaction details (Here it is 'IndusInd_Bank@indusind.com' )
    query = "IndusInd_Bank@indusind.com "

    sheet.cell(row_no, column=1).value = str(date.today())
    book.save("funds2.xlsx")

    while True:
        
        # Retrieving messages from a given sender
        results = service.users().messages().list(userId='me',labelIds = ['INBOX'] , q = query).execute()
        messages = results.get('messages', [])

        # Getting the present time
        
        # To check if it is the next day and initializing the expenses for the next day
        today = date.today()
        
# Send a POST request to the SheetDB API endpoint with the data
        if today not in dates:
            dates.append(today)
            row_no += 1
            lunch_expense = 0
            dinner_expense = 0 
            breakfast_expense = 0
            new_row_data = {
                'date': f"{today}",
                'brkfst': '00',
                'lunch': '00',
                'dinner':'00'
            }
            response = requests.post(URL, json=new_row_data)
            sheet.cell(row_no, column=1).value = str(today)
            book.save("funds2.xlsx")

        if not messages:
            pass
        else:
            
            for message in messages:
                
                msg = service.users().messages().get(userId='me', id=message['id']).execute()
                body = msg['snippet']
                splittd_msg = body.split(" ")
                # Fetching index of the expense from the list containing the splitted string
               
                # Checking whether if keywords ('brea','lun','dinn') are present in the message and updating corresponding
                # parameters in the excel sheet
                try:
                    lis = ['brea' , 'lun' , 'dinn']
                    if 'brea' or 'lun' or 'dinn' in msg['snippet']:
                        print("hello")

                        r = requests.get(url = URL)
                        data = r.json()
                        rows = len(data) + 1
                        indeX = splittd_msg.index("INR") +1
                        last_row = response.json()[-1]
                        if 'brea' in msg['snippet'] and (msg['snippet'] not in breakfast) :
                            print("brea")
                            breakfast_expense += float(splittd_msg[indeX])
                            last_row_par = {'brkfst': f'{breakfast_expense}'}
                            update_url = f"{URL}/brkfst/{last_row['brkfst']}"
                            update_response = requests.patch(update_url, json=last_row_par)
                            sheet.cell(row_no, column=2).value = breakfast_expense
                            book.save("funds2.xlsx")
                            breakfast.append(msg['snippet'])
                            data[-1]['brkfst'] = breakfast_expense

                        if 'lun' in msg['snippet'] and (msg['snippet'] not in lunch) :
                            lunch_expense += float(splittd_msg[indeX])
                            last_row_par = {'lunch': f'{lunch_expense}'}
                            update_url = f"{URL}/lunch/{last_row['lunch']}"
                            update_response = requests.patch(update_url, json=last_row_par)
                            sheet.cell(row_no, column=3).value = lunch_expense
                            book.save("funds2.xlsx")
                            lunch.append(msg['snippet'])
                            data[-1]['lunch'] = lunch_expense
                        
                        if 'dinn' in msg['snippet'] and (msg['snippet'] not in dinner) :
                            dinner_expense += float(splittd_msg[indeX])
                            last_row_par = {'dinner': f'{breakfast_expense}'}
                            update_url = f"{URL}/dinner/{last_row['dinner']}"
                            update_response = requests.patch(update_url, json=last_row_par)
                            sheet.cell(row_no, column=4).value = dinner_expense
                            book.save("funds2.xlsx")
                            dinner.append(msg['snippet'])
                            data[-1]['dinner'] = dinner_expense
                    
                except:
                    pass
                
# Executing the main function
if __name__ == '__main__':
    main()